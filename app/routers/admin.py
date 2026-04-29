import os, uuid, json
from pathlib import Path
from typing import Optional

from fastapi import (APIRouter, Depends, Request, HTTPException,
                     Form, UploadFile, File, Query)
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete, update
from sqlalchemy.orm import joinedload

from app.models import get_db, Product, Category, Service, BlogPost, Order, Brand
from app.services.auth import check_admin, admin_login_response, require_admin
from app.services.seo import slugify, product_seo, category_seo, service_seo, blog_seo
from app.config import settings

router = APIRouter(prefix="/admin")
templates = Jinja2Templates(directory="app/templates")

UPLOAD_DIR = Path("app/static/uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def admin_ctx(request: Request) -> dict:
    return {"request": request, "site_name": settings.SITE_NAME}


# ─── Auth ──────────────────────────────────────────────────

@router.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, error: str = ""):
    return templates.TemplateResponse("admin/login.html", {"request": request, "error": error})

@router.post("/login")
async def login(request: Request, login: str = Form(...), password: str = Form(...)):
    response = RedirectResponse("/admin/", status_code=302)
    if admin_login_response(response, login, password):
        return response
    return templates.TemplateResponse("admin/login.html", {"request": request, "error": "Неверный логин или пароль"})

@router.get("/logout")
async def logout():
    r = RedirectResponse("/admin/login", status_code=302)
    r.delete_cookie("admin_token")
    return r


# ─── Dashboard ─────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    stats = {
        "products": (await db.execute(select(func.count(Product.id)))).scalar(),
        "categories": (await db.execute(select(func.count(Category.id)))).scalar(),
        "services": (await db.execute(select(func.count(Service.id)))).scalar(),
        "posts": (await db.execute(select(func.count(BlogPost.id)))).scalar(),
        "orders": (await db.execute(select(func.count(Order.id)))).scalar(),
        "new_orders": (await db.execute(select(func.count(Order.id)).where(Order.status == "new"))).scalar(),
        "brands": (await db.execute(select(func.count(Brand.id)))).scalar(),
    }
    recent_orders = (await db.execute(select(Order).order_by(Order.created_at.desc()).limit(10))).scalars().all()
    ctx = admin_ctx(request)
    ctx.update({"stats": stats, "recent_orders": recent_orders})
    return templates.TemplateResponse("admin/dashboard.html", ctx)


# ─── Brands ────────────────────────────────────────────────

@router.get("/brands", response_class=HTMLResponse)
async def brands_list(request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    brands = (await db.execute(select(Brand).order_by(Brand.sort_order, Brand.name))).scalars().all()
    brand_counts = {}
    for b in brands:
        count = (await db.execute(
            select(func.count(Product.id)).where(Product.brand_id == b.id)
        )).scalar()
        brand_counts[b.id] = count or 0
    return templates.TemplateResponse("admin/brands.html", {**admin_ctx(request), "brands": brands, "brand_counts": brand_counts})

@router.get("/brands/new", response_class=HTMLResponse)
async def brand_new(request: Request):
    require_admin(request)
    return templates.TemplateResponse("admin/brand_form.html", {**admin_ctx(request), "brand": None})

@router.post("/brands/new")
async def brand_create(
    request: Request,
    name: str = Form(...), description: str = Form(""),
    website: str = Form(""), sort_order: int = Form(0),
    seo_title: str = Form(""), seo_description: str = Form(""),
    logo: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db)
):
    require_admin(request)
    form = await request.form()
    is_official_dealer = "is_official_dealer" in form
    show_in_footer = "show_in_footer" in form
    show_in_catalog = "show_in_catalog" in form
    logo_url = ""
    if logo and logo.filename:
        logo_url = await save_upload(logo)
    b = Brand(
        name=name, slug=slugify(name), description=description,
        website=website, logo_url=logo_url, sort_order=sort_order,
        is_official_dealer=is_official_dealer,
        show_in_footer=show_in_footer, show_in_catalog=show_in_catalog,
        seo_title=seo_title, seo_description=seo_description,
    )
    db.add(b)
    await db.commit()
    return RedirectResponse("/admin/brands", status_code=302)

@router.get("/brands/{bid}/edit", response_class=HTMLResponse)
async def brand_edit(bid: int, request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    brand = (await db.execute(select(Brand).where(Brand.id == bid))).scalar_one_or_none()
    if not brand: raise HTTPException(404)
    return templates.TemplateResponse("admin/brand_form.html", {**admin_ctx(request), "brand": brand})

@router.post("/brands/{bid}/edit")
async def brand_update(
    bid: int, request: Request,
    name: str = Form(...), description: str = Form(""),
    website: str = Form(""), sort_order: int = Form(0),
    seo_title: str = Form(""), seo_description: str = Form(""),
    logo: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db)
):
    require_admin(request)
    form = await request.form()
    is_official_dealer = "is_official_dealer" in form
    show_in_footer = "show_in_footer" in form
    show_in_catalog = "show_in_catalog" in form
    delete_logo = form.get("delete_logo", "")
    brand = (await db.execute(select(Brand).where(Brand.id == bid))).scalar_one_or_none()
    if not brand: raise HTTPException(404)
    brand.name = name; brand.description = description
    brand.website = website; brand.sort_order = sort_order
    brand.is_official_dealer = is_official_dealer
    brand.show_in_footer = show_in_footer
    brand.show_in_catalog = show_in_catalog
    brand.seo_title = seo_title; brand.seo_description = seo_description
    if delete_logo:
        _delete_file(brand.logo_url); brand.logo_url = ""
    if logo and logo.filename:
        brand.logo_url = await save_upload(logo)
    await db.commit()
    return RedirectResponse("/admin/brands", status_code=302)

@router.post("/brands/{bid}/delete")
async def brand_delete(bid: int, request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    await db.execute(update(Product).where(Product.brand_id == bid).values(brand_id=None))
    await db.execute(delete(Brand).where(Brand.id == bid))
    await db.commit()
    return RedirectResponse("/admin/brands", status_code=302)


# ─── Categories ────────────────────────────────────────────

@router.get("/categories", response_class=HTMLResponse)
async def categories_list(request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    cats = (await db.execute(select(Category).order_by(Category.sort_order))).scalars().all()
    return templates.TemplateResponse("admin/categories.html", {**admin_ctx(request), "categories": cats})

@router.get("/categories/new", response_class=HTMLResponse)
async def category_new(request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    parents = (await db.execute(select(Category).where(Category.parent_id == None).order_by(Category.name))).scalars().all()
    return templates.TemplateResponse("admin/category_form.html", {**admin_ctx(request), "cat": None, "parents": parents})

@router.post("/categories/new")
async def category_create(
    request: Request,
    name: str = Form(...), description: str = Form(""),
    parent_id: Optional[int] = Form(None), sort_order: int = Form(0),
    seo_title: str = Form(""), seo_description: str = Form(""),
    image: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db)
):
    require_admin(request)
    image_url = ""
    if image and image.filename:
        image_url = await save_upload(image)
    cat = Category(name=name, slug=slugify(name), description=description,
                   parent_id=parent_id or None, sort_order=sort_order,
                   seo_title=seo_title, seo_description=seo_description, image_url=image_url)
    db.add(cat)
    await db.commit()
    return RedirectResponse("/admin/categories", status_code=302)

@router.get("/categories/{cat_id}/edit", response_class=HTMLResponse)
async def category_edit(cat_id: int, request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    cat = (await db.execute(select(Category).where(Category.id == cat_id))).scalar_one_or_none()
    if not cat: raise HTTPException(404)
    parents = (await db.execute(select(Category).where(Category.parent_id == None, Category.id != cat_id))).scalars().all()
    return templates.TemplateResponse("admin/category_form.html", {**admin_ctx(request), "cat": cat, "parents": parents})

@router.post("/categories/{cat_id}/edit")
async def category_update(
    cat_id: int, request: Request,
    name: str = Form(...), description: str = Form(""),
    parent_id: Optional[int] = Form(None), sort_order: int = Form(0),
    seo_title: str = Form(""), seo_description: str = Form(""),
    image: Optional[UploadFile] = File(None), delete_image: str = Form(""),
    db: AsyncSession = Depends(get_db)
):
    require_admin(request)
    cat = (await db.execute(select(Category).where(Category.id == cat_id))).scalar_one_or_none()
    if not cat: raise HTTPException(404)
    cat.name = name; cat.description = description
    cat.parent_id = parent_id or None; cat.sort_order = sort_order
    cat.seo_title = seo_title; cat.seo_description = seo_description
    if delete_image:
        _delete_file(cat.image_url); cat.image_url = ""
    if image and image.filename:
        cat.image_url = await save_upload(image)
    await db.commit()
    return RedirectResponse("/admin/categories", status_code=302)

@router.post("/categories/{cat_id}/delete")
async def category_delete(cat_id: int, request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    await db.execute(update(Product).where(Product.category_id == cat_id).values(category_id=None))
    await db.execute(delete(Category).where(Category.id == cat_id))
    await db.commit()
    return RedirectResponse("/admin/categories", status_code=302)


# ─── Products Tree ────────────────────────────────────

@router.get("/products/tree", response_class=HTMLResponse)
async def products_tree(request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    root_cats = (await db.execute(select(Category).where(Category.parent_id == None).order_by(Category.sort_order, Category.name))).scalars().all()
    cat_counts = {}
    for cat in root_cats:
        direct = (await db.execute(select(func.count(Product.id)).where(Product.category_id == cat.id))).scalar() or 0
        subcats = (await db.execute(select(Category).where(Category.parent_id == cat.id))).scalars().all()
        sub_total = sum([(await db.execute(select(func.count(Product.id)).where(Product.category_id == s.id))).scalar() or 0 for s in subcats])
        cat_counts[cat.id] = {"direct": direct, "total": direct + sub_total}
    no_cat_count = (await db.execute(select(func.count(Product.id)).where(Product.category_id == None))).scalar() or 0
    ctx = admin_ctx(request)
    ctx.update({"root_cats": root_cats, "cat_counts": cat_counts, "no_cat_count": no_cat_count})
    return templates.TemplateResponse("admin/products_tree.html", ctx)


@router.get("/products/category/{cat_id}", response_class=HTMLResponse)
async def products_in_category(cat_id: int, request: Request, page: int = Query(1, ge=1), db: AsyncSession = Depends(get_db)):
    require_admin(request)
    cat = (await db.execute(select(Category).where(Category.id == cat_id))).scalar_one_or_none()
    if not cat: raise HTTPException(404)
    breadcrumb = [cat]
    if cat.parent_id:
        parent = (await db.execute(select(Category).where(Category.id == cat.parent_id))).scalar_one_or_none()
        if parent: breadcrumb.insert(0, parent)
    subcats = (await db.execute(select(Category).where(Category.parent_id == cat_id).order_by(Category.sort_order, Category.name))).scalars().all()
    sub_counts = {}
    for sub in subcats:
        sub_counts[sub.id] = (await db.execute(select(func.count(Product.id)).where(Product.category_id == sub.id))).scalar() or 0
    per_page = 30
    offset = (page - 1) * per_page
    products = (await db.execute(select(Product).where(Product.category_id == cat_id).order_by(Product.sort_order, Product.id.desc()).offset(offset).limit(per_page))).scalars().all()
    total = (await db.execute(select(func.count(Product.id)).where(Product.category_id == cat_id))).scalar() or 0
    ctx = admin_ctx(request)
    ctx.update({"cat": cat, "breadcrumb": breadcrumb, "subcats": subcats, "sub_counts": sub_counts, "products": products, "total": total, "page": page, "pages": (total + per_page - 1) // per_page})
    return templates.TemplateResponse("admin/products_category.html", ctx)


@router.get("/products/no-category", response_class=HTMLResponse)
async def products_no_category(request: Request, page: int = Query(1, ge=1), db: AsyncSession = Depends(get_db)):
    require_admin(request)
    per_page = 30
    offset = (page - 1) * per_page
    products = (await db.execute(select(Product).where(Product.category_id == None).order_by(Product.id.desc()).offset(offset).limit(per_page))).scalars().all()
    total = (await db.execute(select(func.count(Product.id)).where(Product.category_id == None))).scalar() or 0
    ctx = admin_ctx(request)
    ctx.update({"products": products, "total": total, "page": page, "pages": (total + per_page - 1) // per_page})
    return templates.TemplateResponse("admin/products_no_category.html", ctx)


# ─── Products ──────────────────────────────────────────────

@router.get("/products", response_class=HTMLResponse)
async def products_list(request: Request, page: int = Query(1, ge=1), q: str = Query(""), db: AsyncSession = Depends(get_db)):
    require_admin(request)
    per_page = 30
    offset = (page - 1) * per_page
    query = select(Product)
    if q:
        from sqlalchemy import or_
        like = f"%{q}%"
        query = query.where(or_(Product.name.ilike(like), Product.sku.ilike(like)))
    total = (await db.execute(select(func.count(Product.id)))).scalar()
    products = (await db.execute(query.order_by(Product.id.desc()).offset(offset).limit(per_page))).scalars().all()
    cats = (await db.execute(select(Category).order_by(Category.name))).scalars().all()
    brands = (await db.execute(select(Brand).order_by(Brand.name))).scalars().all()
    return templates.TemplateResponse("admin/products.html", {
        **admin_ctx(request), "products": products, "total": total,
        "page": page, "pages": (total + per_page - 1) // per_page, "q": q,
        "categories": cats, "brands": brands
    })

@router.get("/products/new", response_class=HTMLResponse)
async def product_new(request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    cats = (await db.execute(select(Category).order_by(Category.name))).scalars().all()
    brands = (await db.execute(select(Brand).where(Brand.show_in_catalog == True).order_by(Brand.name))).scalars().all()
    return templates.TemplateResponse("admin/product_form.html", {
        **admin_ctx(request), "product": None, "categories": cats, "brands": brands, "auto_seo": {}
    })

@router.post("/products/new")
async def product_create(
    request: Request,
    name: str = Form(...), sku: str = Form(""), brand: str = Form(""),
    brand_id: Optional[int] = Form(None),
    category_id: Optional[int] = Form(None),
    description: str = Form(""), short_description: str = Form(""),
    price: float = Form(0), old_price: float = Form(0),
    in_stock: bool = Form(True), is_active: bool = Form(True), is_featured: bool = Form(False),
    video_url: str = Form(""), specs_json: str = Form("{}"),
    seo_title: str = Form(""), seo_description: str = Form(""),
    sort_order: int = Form(0),
    main_image: Optional[UploadFile] = File(None),
    extra_images: list[UploadFile] = File(default=[]),
    db: AsyncSession = Depends(get_db)
):
    require_admin(request)
    slug = slugify(name)
    existing = (await db.execute(select(Product).where(Product.slug == slug))).scalar_one_or_none()
    if existing:
        slug = f"{slug}-{sku.lower() or uuid.uuid4().hex[:6]}"

    # images[0] = основное, остальные — дополнительные
    saved_images = []
    if main_image and main_image.filename:
        saved_images.append(await save_upload(main_image))
    for img in extra_images:
        if img and img.filename:
            saved_images.append(await save_upload(img))

    try:
        specs = json.loads(specs_json)
    except Exception:
        specs = {}

    brand_name = brand
    if brand_id:
        b = (await db.execute(select(Brand).where(Brand.id == brand_id))).scalar_one_or_none()
        if b:
            brand_name = b.name

    p = Product(
        name=name, slug=slug, sku=sku, brand=brand_name, brand_id=brand_id or None,
        category_id=category_id or None,
        description=description, short_description=short_description,
        price=price, old_price=old_price, in_stock=in_stock,
        is_active=is_active, is_featured=is_featured,
        images=saved_images, video_url=video_url, specs=specs, sort_order=sort_order,
    )
    db.add(p)
    await db.commit()
    await db.refresh(p)
    p.seo_title = seo_title if seo_title.strip() else product_seo(p)["title"]
    p.seo_description = seo_description if seo_description.strip() else product_seo(p)["description"]
    await db.commit()
    return RedirectResponse("/admin/products", status_code=302)

@router.get("/products/{pid}/edit", response_class=HTMLResponse)
async def product_edit(pid: int, request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    product = (await db.execute(select(Product).where(Product.id == pid))).scalar_one_or_none()
    if not product: raise HTTPException(404)
    cats = (await db.execute(select(Category).order_by(Category.name))).scalars().all()
    brands = (await db.execute(select(Brand).where(Brand.show_in_catalog == True).order_by(Brand.name))).scalars().all()
    auto_seo = product_seo(product)
    return templates.TemplateResponse("admin/product_form.html", {
        **admin_ctx(request), "product": product, "categories": cats, "brands": brands, "auto_seo": auto_seo
    })

@router.post("/products/{pid}/edit")
async def product_update(
    pid: int, request: Request,
    name: str = Form(...), sku: str = Form(""), brand: str = Form(""),
    brand_id: Optional[int] = Form(None),
    category_id: Optional[int] = Form(None),
    description: str = Form(""), short_description: str = Form(""),
    price: float = Form(0), old_price: float = Form(0),
    in_stock: bool = Form(True), is_active: bool = Form(True), is_featured: bool = Form(False),
    video_url: str = Form(""), specs_json: str = Form("{}"),
    seo_title: str = Form(""), seo_description: str = Form(""),
    sort_order: int = Form(0),
    main_image: Optional[UploadFile] = File(None),
    extra_images: list[UploadFile] = File(default=[]),
    delete_images: str = Form(""),
    db: AsyncSession = Depends(get_db)
):
    require_admin(request)
    p = (await db.execute(select(Product).where(Product.id == pid))).scalar_one_or_none()
    if not p: raise HTTPException(404)

    try:
        specs = json.loads(specs_json)
    except Exception:
        specs = {}

    current_images = list(p.images or [])

    # Удаляем отмеченные фото
    if delete_images:
        for img_path in delete_images.split(","):
            img_path = img_path.strip()
            if img_path and img_path in current_images:
                current_images.remove(img_path)
                _delete_file(img_path)

    # Новое основное фото — вставляем на позицию 0
    if main_image and main_image.filename:
        new_main = await save_upload(main_image)
        current_images.insert(0, new_main)

    # Дополнительные — добавляем в конец
    for img in extra_images:
        if img and img.filename:
            current_images.append(await save_upload(img))

    brand_name = brand
    if brand_id:
        b = (await db.execute(select(Brand).where(Brand.id == brand_id))).scalar_one_or_none()
        if b:
            brand_name = b.name

    p.name = name; p.sku = sku; p.brand = brand_name; p.brand_id = brand_id or None
    p.category_id = category_id or None
    p.description = description; p.short_description = short_description
    p.price = price; p.old_price = old_price
    p.in_stock = in_stock; p.is_active = is_active; p.is_featured = is_featured
    p.images = current_images; p.video_url = video_url; p.specs = specs; p.sort_order = sort_order
    p.seo_title = seo_title if seo_title.strip() else product_seo(p)["title"]
    p.seo_description = seo_description if seo_description.strip() else product_seo(p)["description"]
    await db.commit()
    return RedirectResponse("/admin/products", status_code=302)

@router.post("/products/{pid}/delete")
async def product_delete(pid: int, request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    await db.execute(delete(Product).where(Product.id == pid))
    await db.commit()
    return RedirectResponse("/admin/products", status_code=302)


# ─── Services ──────────────────────────────────────────────

@router.get("/services", response_class=HTMLResponse)
async def services_list(request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    svcs = (await db.execute(select(Service).order_by(Service.sort_order))).scalars().all()
    return templates.TemplateResponse("admin/services.html", {**admin_ctx(request), "services": svcs})

@router.get("/services/new", response_class=HTMLResponse)
async def service_new(request: Request):
    require_admin(request)
    return templates.TemplateResponse("admin/service_form.html", {**admin_ctx(request), "service": None, "auto_seo": {}})

@router.post("/services/new")
async def service_create(
    request: Request,
    name: str = Form(...), description: str = Form(""), short_description: str = Form(""),
    price_from: float = Form(0), price_to: float = Form(0),
    icon: str = Form("shield"), is_active: bool = Form(True), is_featured: bool = Form(False),
    seo_title: str = Form(""), seo_description: str = Form(""),
    sort_order: int = Form(0),
    image: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db)
):
    require_admin(request)
    image_url = ""
    if image and image.filename:
        image_url = await save_upload(image)
    svc = Service(name=name, slug=slugify(name), description=description,
                  short_description=short_description, price_from=price_from, price_to=price_to,
                  icon=icon, is_active=is_active, is_featured=is_featured,
                  image_url=image_url, sort_order=sort_order)
    db.add(svc)
    await db.commit()
    await db.refresh(svc)
    svc.seo_title = seo_title if seo_title.strip() else service_seo(svc)["title"]
    svc.seo_description = seo_description if seo_description.strip() else service_seo(svc)["description"]
    await db.commit()
    return RedirectResponse("/admin/services", status_code=302)

@router.get("/services/{sid}/edit", response_class=HTMLResponse)
async def service_edit(sid: int, request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    svc = (await db.execute(select(Service).where(Service.id == sid))).scalar_one_or_none()
    if not svc: raise HTTPException(404)
    return templates.TemplateResponse("admin/service_form.html", {
        **admin_ctx(request), "service": svc, "auto_seo": service_seo(svc)
    })

@router.post("/services/{sid}/edit")
async def service_update(
    sid: int, request: Request,
    name: str = Form(...), description: str = Form(""), short_description: str = Form(""),
    price_from: float = Form(0), price_to: float = Form(0),
    icon: str = Form("shield"), is_active: bool = Form(True), is_featured: bool = Form(False),
    seo_title: str = Form(""), seo_description: str = Form(""),
    sort_order: int = Form(0),
    image: Optional[UploadFile] = File(None), delete_image: str = Form(""),
    db: AsyncSession = Depends(get_db)
):
    require_admin(request)
    svc = (await db.execute(select(Service).where(Service.id == sid))).scalar_one_or_none()
    if not svc: raise HTTPException(404)
    svc.name = name; svc.description = description; svc.short_description = short_description
    svc.price_from = price_from; svc.price_to = price_to; svc.icon = icon
    svc.is_active = is_active; svc.is_featured = is_featured; svc.sort_order = sort_order
    if delete_image:
        _delete_file(svc.image_url); svc.image_url = ""
    if image and image.filename:
        svc.image_url = await save_upload(image)
    svc.seo_title = seo_title if seo_title.strip() else service_seo(svc)["title"]
    svc.seo_description = seo_description if seo_description.strip() else service_seo(svc)["description"]
    await db.commit()
    return RedirectResponse("/admin/services", status_code=302)

@router.post("/services/{sid}/delete")
async def service_delete(sid: int, request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    await db.execute(delete(Service).where(Service.id == sid))
    await db.commit()
    return RedirectResponse("/admin/services", status_code=302)


# ─── Blog ──────────────────────────────────────────────────

@router.get("/blog", response_class=HTMLResponse)
async def blog_list(request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    posts = (await db.execute(select(BlogPost).order_by(BlogPost.created_at.desc()))).scalars().all()
    return templates.TemplateResponse("admin/blog.html", {**admin_ctx(request), "posts": posts})

@router.get("/blog/new", response_class=HTMLResponse)
async def blog_new(request: Request):
    require_admin(request)
    return templates.TemplateResponse("admin/blog_form.html", {**admin_ctx(request), "post": None})

@router.post("/blog/new")
async def blog_create(
    request: Request,
    title: str = Form(...), excerpt: str = Form(""), content: str = Form(""),
    is_published: bool = Form(False),
    seo_title: str = Form(""), seo_description: str = Form(""), seo_keywords: str = Form(""),
    cover: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db)
):
    require_admin(request)
    cover_image = ""
    if cover and cover.filename:
        cover_image = await save_upload(cover)
    post = BlogPost(title=title, slug=slugify(title), excerpt=excerpt, content=content,
                    is_published=is_published, cover_image=cover_image, seo_keywords=seo_keywords)
    db.add(post)
    await db.commit()
    await db.refresh(post)
    post.seo_title = seo_title if seo_title.strip() else blog_seo(post)["title"]
    post.seo_description = seo_description if seo_description.strip() else blog_seo(post)["description"]
    await db.commit()
    return RedirectResponse("/admin/blog", status_code=302)

@router.get("/blog/{pid}/edit", response_class=HTMLResponse)
async def blog_edit(pid: int, request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    post = (await db.execute(select(BlogPost).where(BlogPost.id == pid))).scalar_one_or_none()
    if not post: raise HTTPException(404)
    return templates.TemplateResponse("admin/blog_form.html", {**admin_ctx(request), "post": post})

@router.post("/blog/{pid}/edit")
async def blog_update(
    pid: int, request: Request,
    title: str = Form(...), excerpt: str = Form(""), content: str = Form(""),
    is_published: bool = Form(False),
    seo_title: str = Form(""), seo_description: str = Form(""), seo_keywords: str = Form(""),
    cover: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db)
):
    require_admin(request)
    post = (await db.execute(select(BlogPost).where(BlogPost.id == pid))).scalar_one_or_none()
    if not post: raise HTTPException(404)
    post.title = title; post.excerpt = excerpt; post.content = content
    post.is_published = is_published; post.seo_keywords = seo_keywords
    if cover and cover.filename:
        post.cover_image = await save_upload(cover)
    post.seo_title = seo_title if seo_title.strip() else blog_seo(post)["title"]
    post.seo_description = seo_description if seo_description.strip() else blog_seo(post)["description"]
    await db.commit()
    return RedirectResponse("/admin/blog", status_code=302)

@router.post("/blog/{pid}/delete")
async def blog_delete(pid: int, request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    await db.execute(delete(BlogPost).where(BlogPost.id == pid))
    await db.commit()
    return RedirectResponse("/admin/blog", status_code=302)


# ─── Orders ────────────────────────────────────────────────

@router.get("/orders", response_class=HTMLResponse)
async def orders_list(request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    orders = (await db.execute(select(Order).order_by(Order.created_at.desc()))).scalars().all()
    return templates.TemplateResponse("admin/orders.html", {**admin_ctx(request), "orders": orders})

@router.get("/orders/{oid}", response_class=HTMLResponse)
async def order_detail(oid: int, request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    order = (await db.execute(select(Order).where(Order.id == oid))).scalar_one_or_none()
    if not order: raise HTTPException(404)
    return templates.TemplateResponse("admin/order_detail.html", {**admin_ctx(request), "order": order})

@router.post("/orders/{oid}/status")
async def order_status(oid: int, request: Request, status: str = Form(...), db: AsyncSession = Depends(get_db)):
    require_admin(request)
    order = (await db.execute(select(Order).where(Order.id == oid))).scalar_one_or_none()
    if not order: raise HTTPException(404)
    order.status = status
    await db.commit()
    return RedirectResponse(f"/admin/orders/{oid}", status_code=302)


# ─── Parser ────────────────────────────────────────────────

@router.post("/run-parser")
async def run_parser_now(request: Request):
    require_admin(request)
    from app.services.parser import run_parser
    import asyncio
    asyncio.create_task(run_parser())
    return RedirectResponse("/admin/", status_code=302)


# ─── Helpers ───────────────────────────────────────────────

async def save_upload(file: UploadFile) -> str:
    ext = Path(file.filename).suffix.lower()
    if ext not in {".jpg", ".jpeg", ".png", ".webp", ".gif", ".svg"}:
        raise HTTPException(400, "Недопустимый формат файла")
    fname = f"{uuid.uuid4().hex}{ext}"
    dest = UPLOAD_DIR / fname
    content = await file.read()
    with open(dest, "wb") as f:
        f.write(content)
    return f"/static/uploads/{fname}"

def _delete_file(path: str):
    if not path:
        return
    try:
        full = Path("app") / path.lstrip("/")
        if full.exists():
            full.unlink()
    except Exception:
        pass


@router.post("/products/bulk-edit")
async def products_bulk_edit(
    request: Request,
    product_ids: str = Form(...),
    action_category_id: Optional[int] = Form(None),
    action_brand_id: Optional[int] = Form(None),
    price_action: str = Form(""),
    price_value: float = Form(0),
    db: AsyncSession = Depends(get_db)
):
    require_admin(request)
    ids = [int(x) for x in product_ids.split(",") if x.strip().isdigit()]
    if not ids:
        return RedirectResponse("/admin/products", status_code=302)

    products = (await db.execute(select(Product).where(Product.id.in_(ids)))).scalars().all()

    for p in products:
        if action_category_id:
            p.category_id = action_category_id
        if action_brand_id:
            b = (await db.execute(select(Brand).where(Brand.id == action_brand_id))).scalar_one_or_none()
            if b:
                p.brand_id = action_brand_id
                p.brand = b.name
        if price_action and price_value > 0:
            if price_action == "plus_pct":
                p.old_price = 0
                p.price = round(p.price * (1 + price_value / 100), 2)
            elif price_action == "minus_pct":
                p.old_price = 0
                p.price = round(p.price * (1 - price_value / 100), 2)

    await db.commit()
    return RedirectResponse("/admin/products", status_code=302)


@router.get("/import", response_class=HTMLResponse)
async def import_page(request: Request, db: AsyncSession = Depends(get_db)):
    require_admin(request)
    cats = (await db.execute(select(Category).order_by(Category.name))).scalars().all()
    brands = (await db.execute(select(Brand).order_by(Brand.name))).scalars().all()
    return templates.TemplateResponse("admin/import.html", {
        **admin_ctx(request), "categories": cats, "brands": brands
    })


@router.post("/import/preview")
async def import_preview(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    require_admin(request)
    try:
        import openpyxl, io
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        headers = {}
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        for i, h in enumerate(first_row):
            if h:
                headers[str(h).strip().lower()] = i

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v for v in row if v is not None):
                continue
            def get(key, fallback_idx):
                idx = headers.get(key, fallback_idx)
                return row[idx] if idx < len(row) else None

            name = get('название', 0)
            if not name:
                continue

            rows.append({
                "name": str(name).strip(),
                "short_description": str(get('краткое описание', 2) or '').strip()[:500],
                "price": float(get('цена', 9) or 0),
                "category": str(get('категория', 11) or '').strip(),
                "brand": str(get('бренд', 13) or '').strip(),
            })

        import json, tempfile
        tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, dir='/tmp')
        json.dump(rows, tmp, ensure_ascii=False)
        tmp.close()

        cats = (await db.execute(select(Category).order_by(Category.name))).scalars().all()
        brands_list = (await db.execute(select(Brand).order_by(Brand.name))).scalars().all()

        file_cats = sorted(set(r['category'] for r in rows if r['category']))
        file_brands = sorted(set(r['brand'] for r in rows if r['brand']))

        return templates.TemplateResponse("admin/import_preview.html", {
            **admin_ctx(request),
            "rows": rows[:5],
            "total": len(rows),
            "file_cats": file_cats,
            "file_brands": file_brands,
            "categories": cats,
            "brands_list": brands_list,
            "tmp_file": tmp.name,
        })
    except Exception as e:
        return templates.TemplateResponse("admin/import.html", {
            **admin_ctx(request),
            "error": f"Ошибка чтения файла: {e}",
            "categories": [], "brands": []
        })


@router.post("/import/execute")
async def import_execute(
    request: Request,
    tmp_file: str = Form(...),
    db: AsyncSession = Depends(get_db)
):
    require_admin(request)
    import json
    from app.services.seo import slugify, product_seo

    try:
        with open(tmp_file) as f:
            rows = json.load(f)
    except Exception as e:
        raise HTTPException(400, f"Ошибка: {e}")

    form = await request.form()

    cat_map = {}
    brand_map = {}

    for key, val in form.items():
        if key.startswith("cat_map_") and val:
            file_cat = key[len("cat_map_"):]
            cat_map[file_cat] = int(val)
        if key.startswith("brand_map_") and val:
            file_brand = key[len("brand_map_"):]
            brand_map[file_brand] = int(val)

    created = 0
    updated = 0
    errors = 0

    for row in rows:
        try:
            name = row['name']
            slug = slugify(name)

            existing = (await db.execute(
                select(Product).where(Product.slug == slug)
            )).scalar_one_or_none()

            category_id = cat_map.get(row['category'])
            brand_id = brand_map.get(row['brand'])
            brand_name = row['brand']

            if brand_id:
                b = (await db.execute(select(Brand).where(Brand.id == brand_id))).scalar_one_or_none()
                if b:
                    brand_name = b.name

            if existing:
                existing.short_description = row['short_description']
                existing.price = row['price']
                if category_id:
                    existing.category_id = category_id
                if brand_id:
                    existing.brand_id = brand_id
                    existing.brand = brand_name
                elif row['brand']:
                    existing.brand = row['brand']
                updated += 1
            else:
                check = (await db.execute(select(Product).where(Product.slug == slug))).scalar_one_or_none()
                if check:
                    slug = f"{slug}-{created}"

                p = Product(
                    name=name, slug=slug,
                    short_description=row['short_description'],
                    price=row['price'],
                    category_id=category_id,
                    brand_id=brand_id,
                    brand=brand_name,
                    is_active=True,
                )
                db.add(p)
                await db.flush()
                p.seo_title = product_seo(p)["title"]
                p.seo_description = product_seo(p)["description"]
                created += 1
        except Exception as e:
            errors += 1
            import logging
            logging.getLogger(__name__).error(f"Import error row {row.get('name')}: {e}")

    await db.commit()

    import os
    try:
        os.unlink(tmp_file)
    except Exception:
        pass

    return templates.TemplateResponse("admin/import_result.html", {
        **admin_ctx(request),
        "created": created, "updated": updated, "errors": errors
    })


@router.post("/upload-image")
async def upload_image(request: Request, file: UploadFile = File(...)):
    require_admin(request)
    try:
        url = await save_upload(file)
        return {"url": url}
    except Exception as e:
        return {"error": str(e)}